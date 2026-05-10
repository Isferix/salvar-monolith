from pathlib import Path
from typing import List

from core.pydantic.common import (
    FamiliarCargado,
    FamiliarDesconocido,
    FamiliarRelacion,
    Ubicacion,
)
from core.pydantic.models import GrupoFamiliar, Persona
from core.pydantic.ports import PersonasSerializer, name_resolver, saver
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, Side


class ExcelSerializer(PersonasSerializer):
    def __init__(self, path: Path):
        self.file_path = path

    def _build_grupo(self, familiares_nombres):
        familiares = []

        for idx, texto in enumerate(familiares_nombres, start=1):
            texto: str = texto.strip()

            # -------------------------
            # Caso desconocido
            # -------------------------
            if texto.lower().startswith("Desconocido"):
                familiar = FamiliarDesconocido(
                    nombre_completo=texto,
                    relacion=FamiliarRelacion.OTRO,
                )
            else:
                # -------------------------
                # Heurística: ¿es DNI?
                # -------------------------
                if texto.upper().startswith("DNI"):
                    # Ej: "DNI 12345678"
                    dni = texto.replace("DNI", "").strip()

                    familiar = FamiliarCargado(
                        dni=dni,
                        relacion=FamiliarRelacion.OTRO,
                    )
                else:
                    # -------------------------
                    # Nombre → desconocido
                    # -------------------------
                    familiar = FamiliarDesconocido(
                        nombre_completo=texto,
                        relacion=FamiliarRelacion.OTRO,
                    )

            familiares.append(familiar)

        grupo = GrupoFamiliar(
            familiares=familiares,
        )

        return grupo

    def import_data(self, path: str, saver: saver) -> None:
        wb: Workbook = load_workbook(path)
        if wb is None:
            raise ValueError(f"No se pudo cargar el archivo Excel: {path}")
        ws = wb.active
        if ws is None:
            raise ValueError(f"No se pudo acceder a la hoja activa del Excel: {path}")

        current_persona = None
        new_id_counter = 1
        familiares_buffer = []

        for row in ws.iter_rows(min_row=2, min_col=2, max_col=8, values_only=True):
            print(f"Procesando fila: {row}")
            (
                apellido,
                nombre,
                dni,
                cantidad_grupo_fliar,
                direccion,
                localidad,
                descripcion,
            ) = row

            # -------------------------
            # Detectar subfila
            # -------------------------
            es_subfila = nombre is not None and (
                apellido is None or apellido.strip() == ""
            )

            if es_subfila:
                if current_persona:
                    familiares_buffer.append(nombre.strip())
                continue

            # -------------------------
            # Flush persona anterior
            # -------------------------
            if current_persona:
                if current_persona.grupo_familiar:
                    grupo = self._build_grupo(familiares_buffer)
                    current_persona.grupo_familiar = grupo
                    current_persona.family_owner = True

                saver(current_persona)

            # -------------------------
            # Crear nueva persona y guardar la anterior
            # -------------------------
            current_persona = Persona(
                id=int(new_id_counter),
                apellido=str(apellido),
                nombre=str(nombre),
                dni=str(dni) if dni else None,
                ubicacion=Ubicacion(
                    direccion=str(direccion) if direccion else "",
                    localidad=str(localidad) if localidad else "",
                ),
                grupo_familiar=GrupoFamiliar(
                    cantidad=int(cantidad_grupo_fliar), familiares=[]
                )
                if cantidad_grupo_fliar
                else None,
                family_owner=True if cantidad_grupo_fliar else False,
                cargado_en_caritas=True if cantidad_grupo_fliar else False,
                descripcion=str(descripcion) if descripcion else "",
            )

            familiares_buffer = []
            new_id_counter += 1

        print("Importación finalizada. Flush de la última persona pendiente...")
        # -------------------------
        # Flush última
        # -------------------------
        if current_persona:
            if current_persona.grupo_familiar:
                grupo = self._build_grupo(familiares_buffer)
                current_persona.grupo_familiar = grupo
                current_persona.family_owner = True

            saver(current_persona)

    # -------------------------
    # Construcción correcta dominio
    # -------------------------
    def export_data(self, personas: List[Persona], name_resolver: name_resolver) -> str:
        wb: Workbook = Workbook()
        ws = wb.active
        ws.title = "Personas"
        if ws is None:
            raise ValueError("No se pudo crear la hoja de Excel")

        bold_font = Font(bold=True)
        thin = Side(style="medium")

        full_border = Border(
            left=thin,
            right=thin,
            top=thin,
            bottom=thin,
        )

        no_vertical_border = Border(
            top=thin,
            bottom=thin,
            left=None,
            right=None,
        )

        sub_border_right = Border(
            top=thin,
            bottom=thin,
            right=thin,
            left=None,
        )

        header_font = Font(name="Arial", size=11, bold=True)
        normal_font = Font(name="Calibri", size=11)
        bold_font = Font(name="Calibri", size=11, bold=True)

        align_right = Alignment(horizontal="right")
        align_center = Alignment(horizontal="center")
        wrap_alignment = Alignment(wrap_text=True)

        # -------------------------
        # Header
        # -------------------------
        headers = [
            "ID",
            "Apellido",
            "Nombre",
            "DNI",
            "Grupo Fliar.",
            "Dirección",
            "Localidad",
            "Descripción",
        ]
        ws.append(headers)
        ws.freeze_panes = "A2"
        ws.freeze_panes = "B2"
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = header_font
            cell.border = full_border
            cell.alignment = Alignment(
                horizontal=cell.alignment.horizontal if cell.alignment else None,
                vertical="center",
                wrap_text=(col == 8),  # solo descripción con wrap
            )

        # -------------------------
        # Body
        # -------------------------
        for persona in personas:
            # Grupo familiar
            grupo_valor = None
            if persona.family_owner:
                if persona.grupo_familiar:
                    grupo_valor = persona.grupo_familiar.cantidad
                else:
                    grupo_valor = 1

            row = [
                persona.id,
                persona.apellido,
                persona.nombre,
                persona.dni,
                grupo_valor,
                persona.ubicacion.direccion,
                persona.ubicacion.localidad,
                persona.descripcion,
            ]

            ws.append(row)
            current_row = ws.max_row

            for col in range(1, 9):
                cell = ws.cell(row=current_row, column=col)
                cell.font = bold_font if persona.family_owner else normal_font
                cell.border = full_border

                if col == 4:  # DNI
                    cell.alignment = align_right
                elif col == 1 or col == 5:  # ID y Grupo familiar
                    cell.alignment = align_center
                elif col == 8:  # Descripción
                    cell.alignment = wrap_alignment

            # Negrita si está cargado en caritas
            if persona.cargado_en_caritas:
                for col in range(1, 9):  # columnas A-H
                    ws.cell(row=current_row, column=col).font = bold_font

            # -------------------------
            # Subfilas (familiares)
            # -------------------------
            if (
                persona.family_owner
                and persona.grupo_familiar
                and persona.grupo_familiar.familiares
            ):
                for idx, familiar in enumerate(
                    persona.grupo_familiar.familiares, start=1
                ):
                    nombre_familiar = name_resolver(familiar)
                    if nombre_familiar == "Desconocido":
                        nombre_familiar += f" {idx}"

                    subrow = [
                        "",  # sin ID
                        "",
                        nombre_familiar,
                        "",
                        "",
                        "",
                        "",
                        "",
                    ]

                    ws.append(subrow)
                    sub_row_index = ws.max_row

                    for col in range(1, 9):
                        cell = ws.cell(row=sub_row_index, column=col)
                        cell.font = normal_font

                        if col == 1:
                            # opcional: mantener borde izquierdo externo
                            cell.border = Border(
                                top=thin,
                                bottom=thin,
                                left=thin,
                            )
                        elif col == 8:
                            # 👈 última columna: cerrar derecha
                            cell.border = sub_border_right
                        else:
                            # interior sin verticales
                            cell.border = no_vertical_border

        # -------------------------
        # Ajuste columnas
        # -------------------------
        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter

            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))

            ws.column_dimensions[col_letter].width = max_length + 2
        ws.column_dimensions["H"].width = 40  # descripción más ancha
        wb.save(self.file_path)
        return str(self.file_path)
